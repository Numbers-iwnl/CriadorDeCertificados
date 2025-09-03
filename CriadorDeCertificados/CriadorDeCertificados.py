#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Gerador de Certificados — Wizard (single file, compact-mode + scrollable left)

Requisitos:
  pip install pymupdf pillow openpyxl
"""

import os, sys, json, datetime, re, subprocess, threading, queue, traceback
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
from tkinter import font as tkfont
from dataclasses import dataclass
from typing import List, Optional, Tuple, Dict
from functools import lru_cache

import fitz  # PyMuPDF
from PIL import Image, ImageTk, ImageFont
from openpyxl import load_workbook

APP_TITLE = "Gerador de Certificados — Wizard"
CALIBRATION_SUFFIX = ".calibration.json"

# ----- Paleta -----
C_BG      = "#eef8f4"
C_PRIMARY = "#17c88b"
C_TEAL    = "#009475"
C_DARK    = "#002927"
C_WARM    = "#514a43"
C_BORDER  = "#d6e7e2"
C_SURFACE = "#ffffff"
C_MUTED   = "#6b6b6b"

# ----- Layout (podem ser reduzidos automaticamente em 'compact mode') -----
SIDEBAR_W       = 560
PREVIEW_MIN_W   = 560
PREVIEW_MIN_H   = 340
BAND_HEIGHT     = 60
BANNER1_MAX_H   = 44
BANNER_SIDE_PAD = 16

BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
ASSETS_DIR = os.path.join(BASE_DIR, "assets")
def asset(*parts: str) -> str: return os.path.join(ASSETS_DIR, *parts)

# ----- Campos -----
FIELD_NAME  = "name"
FIELD_CPF   = "cpf"
FIELD_DATE  = "date"
FIELD_TURMA = "turma"
FIELD_TITLES = {FIELD_NAME:"Nome", FIELD_CPF:"CPF", FIELD_DATE:"Data", FIELD_TURMA:"Turma"}
ALL_FIELDS = [FIELD_NAME, FIELD_CPF, FIELD_DATE, FIELD_TURMA]

# ---------- Utilidades ----------
def log_warn(msg: str):
    try: print(f"[WARN] {msg}", file=sys.stderr, flush=True)
    except Exception: pass

def log_err(msg: str, exc: Exception=None):
    try:
        print(f"[ERROR] {msg}", file=sys.stderr)
        if exc is not None: traceback.print_exc()
        sys.stderr.flush()
    except Exception: pass

def sanitize_filename(name: str) -> str:
    return re.sub(r"[\\/:*?\"<>|]", "_", name).strip().strip(".")

def timestamp() -> str:
    return datetime.datetime.now().strftime("%Y-%m-%d_%Hh%M")

def open_folder(path: str):
    try:
        if os.name == "nt": os.startfile(path)
        elif sys.platform == "darwin": subprocess.Popen(["open", path])
        else: subprocess.Popen(["xdg-open", path])
    except Exception as e:
        log_warn(f"open_folder fail: {e}")

def format_cpf(raw: str) -> str:
    digits = re.sub(r"\D","", raw or "")
    if len(digits) == 11:
        return f"{digits[0:3]}.{digits[3:6]}.{digits[6:9]}-{digits[9:11]}"
    return raw or ""

def ensure_font_file(font_path: str) -> str:
    if not font_path or not os.path.isfile(font_path):
        raise FileNotFoundError("Escolha um arquivo de fonte .ttf ou .otf.")
    ext = os.path.splitext(font_path.lower())[1]
    if ext not in (".ttf", ".otf"):
        raise ValueError("A fonte deve ser .ttf ou .otf.")
    return font_path

def unique_path(path: str) -> str:
    """Evita overwrite: 'file.pdf' -> 'file (2).pdf', 'file (3).pdf', ..."""
    if not os.path.exists(path): return path
    base, ext = os.path.splitext(path); i = 2
    while True:
        cand = f"{base} ({i}){ext}"
        if not os.path.exists(cand): return cand
        i += 1

# ---------- XLSX ----------
def read_records_from_xlsx(path: str) -> List[Dict[str,str]]:
    """
    XLSX: Coluna A = Nome (obrigatória)
          Coluna B = CPF  (opcional)
          Coluna C = Turma (opcional)
          Coluna D = Data  (opcional; se ausente → vazio)
    Sem cabeçalho; lê a planilha ativa.
    """
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    records: List[Dict[str,str]] = []
    for row in ws.iter_rows(values_only=True):
        a = (row[0] if len(row)>=1 else None)
        if a is None or str(a).strip()=="": continue
        name  = str(a).strip()
        cpf   = str(row[1]).strip() if (len(row)>=2 and row[1] is not None) else ""
        turma = str(row[2]).strip() if (len(row)>=3 and row[2] is not None) else ""
        datev = row[3] if (len(row)>=4) else None
        if isinstance(datev, datetime.date):
            date_str = datev.strftime("%d/%m/%Y")
        elif datev is None or str(datev).strip()=="":
            date_str = ""
        else:
            date_str = str(datev).strip()
        records.append({FIELD_NAME:name, FIELD_CPF:cpf, FIELD_TURMA:turma, FIELD_DATE:date_str})
    wb.close()
    return records

# ---------- Métricas / Texto ----------
@lru_cache(maxsize=512)
def _pil_font(font_path: str, size_px: int) -> ImageFont.FreeTypeFont:
    return ImageFont.truetype(font_path, size_px)

def _advance_width(text: str, font_path: str, size_pt: float) -> float:
    size_px = max(1, int(round(size_pt)))
    f = _pil_font(font_path, size_px)
    if hasattr(f, "getlength"): return float(f.getlength(text))
    l, t, r, b = f.getbbox(text); return float(r - l)

def _advance_width_pdf(page: fitz.Page, text: str, ttf_path: str, size_pt: float) -> float:
    try:
        return float(page.get_text_length(text, fontsize=size_pt, fontfile=ttf_path))
    except Exception:
        return _advance_width(text, ttf_path, size_pt)

def _font_metrics_em(ttf_path: str) -> Tuple[float, float]:
    try:
        f = fitz.Font(file=ttf_path)
        return float(f.ascender), float(f.descender)
    except Exception: pass
    try:
        size_px = 100
        f = _pil_font(ttf_path, size_px)
        a, d = f.getmetrics()
        total = max(1.0, float(a + d))
        return float(a) / total, -float(d) / total
    except Exception:
        return 0.9, -0.2

def _fits_in_rect(text: str, rect: fitz.Rect, fontfile: str, size_pt: float) -> bool:
    pad_w = rect.width * 0.02
    pad_h = rect.height * 0.06
    target_w = rect.width - 2 * pad_w
    target_h = rect.height - 2 * pad_h
    w = _advance_width(text, fontfile, size_pt)
    asc_em, desc_em = _font_metrics_em(fontfile)
    line_h_pt = (asc_em - desc_em) * size_pt
    return (w <= target_w) and (line_h_pt <= target_h)

def autosize_font_to_rect(text: str, rect: fitz.Rect, fontfile: str,
                          min_size=8.0, max_size=200.0) -> float:
    if not text: return 24.0
    lo, hi = float(min_size), float(max_size); best = lo
    while lo <= hi:
        mid = (lo + hi) / 2.0
        if _fits_in_rect(text, rect, fontfile, mid):
            best = mid; lo = mid + 0.5
        else:
            hi = mid - 0.5
    return max(min_size, min(best, max_size))

def common_font_size_for_all(names: List[str], rect: fitz.Rect, fontfile: str,
                             min_size=8.0, max_size=200.0) -> float:
    if not names: return 24.0
    lo, hi = float(min_size), float(max_size); best = lo
    while lo <= hi:
        mid = (lo + hi) / 2.0
        if all(_fits_in_rect(n, rect, fontfile, mid) for n in names):
            best = mid; lo = mid + 0.5
        else:
            hi = mid - 0.5
    return max(min_size, min(best, max_size))

# --- helpers de desenho com “negrito simulado” (para TURMA/CPF) ---
def _draw_text_repeated(page: fitz.Page, pos: Tuple[float,float], text: str,
                        fontsize: float, fontfile: str, color=(0,0,0),
                        bold: bool=False):
    x, y = pos
    if not bold:
        page.insert_text((x, y), text, fontsize=fontsize, fontfile=fontfile,
                         fontname="UserFont", color=color); return
    offsets = [(0,0), (0.35,0), (0,0.35), (-0.35,0), (0,-0.35)]
    for dx, dy in offsets:
        page.insert_text((x+dx, y+dy), text, fontsize=fontsize, fontfile=fontfile,
                         fontname="UserFont", color=color)

def draw_text_centered(page: fitz.Page, rect: fitz.Rect, text: str,
                       ttf_path: str, fontsize: float, baseline_tweak_ems: float = 0.0,
                       color=(0,0,0), bold: bool=False):
    w = _advance_width_pdf(page, text, ttf_path, fontsize)
    cx = rect.x0 + rect.width / 2.0
    x = cx - w / 2.0
    asc_em, desc_em = _font_metrics_em(ttf_path)
    cy = rect.y0 + rect.height / 2.0
    y_baseline = cy + (asc_em + desc_em) * fontsize / 2.0 + baseline_tweak_ems * fontsize
    _draw_text_repeated(page, (x, y_baseline), text, fontsize, ttf_path, color, bold)

def draw_text_left(page: fitz.Page, rect: fitz.Rect, text: str,
                   ttf_path: str, fontsize: float, baseline_tweak_ems: float = 0.0,
                   color=(0,0,0), bold: bool=False):
    asc_em, desc_em = _font_metrics_em(ttf_path)
    cy = rect.y0 + rect.height / 2.0
    y_baseline = cy + (asc_em + desc_em) * fontsize / 2.0 + baseline_tweak_ems * fontsize
    x = rect.x0 + rect.width * 0.02
    _draw_text_repeated(page, (x, y_baseline), text, fontsize, ttf_path, color, bold)

# ---------- Detecção de áreas ----------
PLACEHOLDER_VARIANTS = ["(Seu Nome Aqui)","Seu Nome Aqui","( Seu Nome Aqui )","(NOME)","NOME","Nome","(Nome)"]

def _get_text_blocks(page: fitz.Page):
    try: blocks = page.get_text("blocks") or []
    except Exception: blocks = []
    rects = []
    for b in blocks:
        if len(b) >= 5 and isinstance(b[4], str) and b[4].strip():
            x0, y0, x1, y1 = b[:4]
            rects.append(fitz.Rect(x0, y0, x1, y1))
    return rects

def _get_word_rects(page: fitz.Page):
    try: words = page.get_text("words") or []
    except Exception: words = []
    rects = []
    for w in words:
        try:
            x0, y0, x1, y1 = w[0], w[1], w[2], w[3]
            rects.append(fitz.Rect(x0, y0, x1, y1))
        except Exception: pass
    return rects

def _get_line_guides(page: fitz.Page) -> Tuple[List[float], List[float]]:
    xs: List[float] = []; ys: List[float] = []
    try:
        drawings = page.get_drawings()
        for item in drawings:
            for it in item.get("items", []):
                if it[0] == "line":
                    p1, p2 = it[1], it[2]
                    x0,y0 = p1; x1,y1 = p2
                    if abs(y1 - y0) < 1.0: ys.append(y0)
                    if abs(x1 - x0) < 1.0: xs.append(x0)
                elif it[0] == "rect":
                    r = it[1]
                    xs.extend([r.x0, r.x1, (r.x0+r.x1)/2])
                    ys.extend([r.y0, r.y1, (r.y0+r.y1)/2])
    except Exception: pass
    return xs, ys

def _find_placeholder_rect(page: fitz.Page) -> Optional[fitz.Rect]:
    for text in PLACEHOLDER_VARIANTS:
        try: rects = page.search_for(text, quads=False)
        except Exception: rects = []
        if rects: return max(rects, key=lambda r: r.get_area())
    return None

def _default_center_rect(page: fitz.Page) -> fitz.Rect:
    pw, ph = page.rect.width, page.rect.height
    w = pw * 0.60; h = ph * 0.10
    cx, cy = pw / 2.0, ph * 0.55
    return fitz.Rect(cx - w/2, cy - h/2, cx + w/2, cy + h/2)

def compute_auto_area_firstpage(doc: fitz.Document) -> fitz.Rect:
    page = doc[0]
    r = _find_placeholder_rect(page)
    if r:
        pad_x = r.width * 0.4; pad_y = r.height * 0.9
        rr = fitz.Rect(r.x0 - pad_x, r.y0 - pad_y, r.x1 + pad_x, r.y1 + pad_y)
        rr.intersect(page.rect); return rr
    return _default_center_rect(page)

def compute_rect_right_of(anchor: fitz.Rect, page: fitz.Page, width_ratio: float, height_mult: float = 1.3) -> fitz.Rect:
    pw = page.rect.width
    h = max(anchor.height * height_mult, pw * 0.02)
    w = max(pw * width_ratio, anchor.width * 2.0)
    x0 = min(page.rect.x1 - w - 4, anchor.x1 + anchor.height * 0.3)
    y0 = anchor.y0 - (h - anchor.height) / 2.0
    r = fitz.Rect(x0, y0, x0 + w, y0 + h)
    r.intersect(page.rect); return r

def detect_field_rects_firstpage(doc: fitz.Document) -> Dict[str, fitz.Rect]:
    result: Dict[str, fitz.Rect] = {}
    page = doc[0]

    try:
        rr = compute_auto_area_firstpage(doc); result[FIELD_NAME] = rr
    except Exception: pass

    def search_one(words: List[str]) -> Optional[fitz.Rect]:
        for w in words:
            try: rects = page.search_for(w, quads=False) or []
            except Exception: rects = []
            if rects: return max(rects, key=lambda a: a.get_area())
        return None

    hit = search_one(["CPF","Cpf","cpf"])
    if hit:
        try: result[FIELD_CPF] = compute_rect_right_of(hit, page, width_ratio=0.35, height_mult=1.25)
        except Exception: pass

    hit = search_one(["TURMA","Turma","turma"])
    if hit:
        try: result[FIELD_TURMA] = compute_rect_right_of(hit, page, width_ratio=0.18, height_mult=1.2)
        except Exception: pass

    hit = search_one(["DATA","Data","data"])
    if hit:
        try: result[FIELD_DATE] = compute_rect_right_of(hit, page, width_ratio=0.22, height_mult=1.2)
        except Exception: pass
    else:
        try:
            words = page.get_text("words") or []
            slashes = [fitz.Rect(w[0],w[1],w[2],w[3]) for w in words if "/" in (w[4] or "")]
            if len(slashes) >= 2:
                slashes.sort(key=lambda r: r.y0)
                for j in range(len(slashes)-1):
                    if abs(slashes[j+1].y0 - slashes[j].y0) < page.rect.height * 0.02:
                        band_y = (slashes[j].y0 + slashes[j+1].y0)/2
                        same = [s for s in slashes if abs(s.y0 - band_y) < page.rect.height*0.02]
                        if len(same) >= 2:
                            r = same[0]; r2 = same[-1]
                            x0 = max(page.rect.x0, r.x0 - page.rect.width*0.05)
                            x1 = min(page.rect.x1, r2.x1 + page.rect.width*0.15)
                            h  = (r2.y1 - r.y0) * 2.0 or page.rect.height*0.03
                            y0 = max(page.rect.y0, r.y0 - h*0.25)
                            result[FIELD_DATE] = fitz.Rect(x0,y0,x1,y0+h); break
        except Exception: pass

    return result

# ---- Detecção de estilo (negrito) perto do rótulo “TURMA” ----
_BOLD_TOKENS = ("bold","black","heavy","semibold","demi","extrabold","ultrabold","medium")

def _detect_style_for_label(page: fitz.Page, label_text: str) -> Dict[str, bool]:
    style = {"bold": False, "italic": False}
    try:
        d = page.get_text("dict"); label_lc = label_text.lower()
        for block in d.get("blocks", []):
            for line in block.get("lines", []):
                for span in line in line.get("spans", []):  # type: ignore
                    pass  # guard for older PyMuPDF? (kept compatibility)
    except Exception:
        pass  # fallback below

    try:
        d = page.get_text("dict"); label_lc = label_text.lower()
        for block in d.get("blocks", []):
            for line in block.get("lines", []):
                for span in line.get("spans", []):
                    txt = (span.get("text") or "").strip()
                    if not txt: continue
                    if label_lc in txt.lower():
                        fname = (span.get("font") or "").lower()
                        if any(tok in fname for tok in _BOLD_TOKENS): style["bold"] = True
                        return style
    except Exception: pass
    return style

def detect_field_styles_firstpage(doc: fitz.Document) -> Dict[str, Dict[str,bool]]:
    page = doc[0]
    styles: Dict[str, Dict[str,bool]] = {}
    styles[FIELD_TURMA] = _detect_style_for_label(page, "TURMA")
    return styles

# ---------- Configuração por template ----------
@dataclass
class FieldCalibration:
    page_index: int
    rect: Tuple[float, float, float, float]
    style: Optional[Dict[str,bool]] = None  # ex.: {"bold": True}

@dataclass
class TemplateConfig:
    required: List[str]
    fields: Dict[str, FieldCalibration]
    version: int = 3  # v3 inclui 'style'

def _load_config(meta_path: str) -> Optional[TemplateConfig]:
    try:
        with open(meta_path, "r", encoding="utf-8") as f: d = json.load(f)
        if "fields" in d and "required" in d:
            fields: Dict[str, FieldCalibration] = {}
            for k, v in d["fields"].items():
                style = v.get("style") if isinstance(v, dict) else None
                fields[k] = FieldCalibration(int(v["page_index"]), tuple(v["rect"]), style=style)
            req = list(d.get("required", [FIELD_NAME])); ver = int(d.get("version", 2))
            return TemplateConfig(required=req, fields=fields, version=ver)
        if "page_index" in d and "rect" in d:
            fc = FieldCalibration(int(d["page_index"]), tuple(d["rect"]), style=None)
            return TemplateConfig(required=[FIELD_NAME], fields={FIELD_NAME: fc}, version=1)
    except Exception as e:
        log_warn(f"_load_config: {e}"); return None
    return None

def load_template_config(template_pdf: str) -> Optional[TemplateConfig]:
    meta = template_pdf + CALIBRATION_SUFFIX
    if os.path.exists(meta): return _load_config(meta)
    return None

def save_template_config(template_pdf: str, cfg: TemplateConfig) -> None:
    meta = template_pdf + CALIBRATION_SUFFIX
    data = {
        "version": cfg.version,
        "required": cfg.required,
        "fields": {
            k: {"page_index": v.page_index, "rect": list(v.rect), "style": (v.style or {})}
            for k, v in cfg.fields.items()
        }
    }
    with open(meta, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

# ---------- Snap ----------
def _collect_snap_candidates(page: fitz.Page) -> Tuple[List[fitz.Rect], List[float], List[float]]:
    rects = _get_text_blocks(page) + _get_word_rects(page)
    xs, ys = _get_line_guides(page)
    pw, ph = page.rect.width, page.rect.height
    grid_x = [pw*0.05, pw*0.10, pw*0.25, pw*0.33, pw*0.5, pw*0.66, pw*0.75, pw*0.90, pw*0.95]
    grid_y = [ph*0.10, ph*0.20, ph*0.33, ph*0.50, ph*0.66, ph*0.80, ph*0.90]
    xs = list(xs) + grid_x + [pw/2]; ys = list(ys) + grid_y + [ph/2]
    return rects, xs, ys

def compute_snapped_rect(page: fitz.Page, rect: fitz.Rect, snap_enabled: bool,
                         tol_pt: float, offset_x: float, offset_y: float) -> fitz.Rect:
    w, h = rect.width, rect.height
    cx, cy = rect.x0 + w/2, rect.y0 + h/2
    if not snap_enabled:
        return fitz.Rect(cx - w/2 + offset_x, cy - h/2 + offset_y,
                         cx + w/2 + offset_x, cy + h/2 + offset_y)

    rects, xs, ys = _collect_snap_candidates(page)
    cx_cands = [page.rect.width/2]
    for r in rects:
        if abs(((r.y0+r.y1)/2) - cy) <= tol_pt*2: cx_cands.append((r.x0 + r.x1)/2)
    cy_cands = [page.rect.height/2]
    for r in rects:
        if abs(((r.x0+r.x1)/2) - cx) <= tol_pt*2: cy_cands.append((r.y0 + r.y1)/2)
    cx_cands += xs; cy_cands += ys

    best_cx = min(cx_cands, key=lambda x: abs(x - cx))
    best_cy = min(cy_cands, key=lambda y: abs(y - cy))

    snapped_cx = cx if abs(best_cx - cx) > tol_pt else best_cx
    snapped_cy = cy if abs(best_cy - cy) > tol_pt else best_cy

    snapped_cx += offset_x; snapped_cy += offset_y
    return fitz.Rect(snapped_cx - w/2, snapped_cy - h/2, snapped_cx + w/2, snapped_cy + h/2)

# ---------- Cor automática ----------
def _avg_luminance_from_pixmap(pix: fitz.Pixmap) -> float:
    try:
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        max_side = 200
        if max(img.size) > max_side:
            ratio = max_side / max(img.size)
            img = img.resize((max(1,int(img.width*ratio)), max(1,int(img.height*ratio))), Image.BILINEAR)
        g = img.convert("L"); hist = g.histogram()
        total = float(sum(hist)) or 1.0
        mean = sum(i*count for i,count in enumerate(hist)) / (255.0*total)
        return float(mean)  # 0=preto, 1=branco
    except Exception as e:
        log_warn(f"_avg_luminance_from_pixmap: {e}"); return 1.0

def pick_text_color(page: fitz.Page, rect: fitz.Rect) -> Tuple[float,float,float]:
    try:
        clip = fitz.Rect(rect.x0, rect.y0, rect.x1, rect.y1)
        pix = page.get_pixmap(matrix=fitz.Matrix(1,1), alpha=False, clip=clip)
        lum = _avg_luminance_from_pixmap(pix)
        return (1,1,1) if lum < 0.40 else (0,0,0)
    except Exception as e:
        log_warn(f"pick_text_color: {e}"); return (0,0,0)

# ---------- UI helpers ----------
class RoundedButton:
    def __init__(self, parent, text, command=None, fill=C_PRIMARY, fg="white", radius=14, padx=22, pady=12):
        self.parent = parent; self.text = text; self.command = command
        self.fill = fill; self.fg = fg; self.radius = radius; self.padx = padx; self.pady = pady
        self.disabled = False
        self.font = tkfont.Font(family="Segoe UI", size=11, weight="bold")
        tw = self.font.measure(self.text); th = self.font.metrics("linespace")
        self.w = tw + self.padx*2; self.h = th + self.pady*2
        self.canvas = tk.Canvas(parent, width=self.w, height=self.h,
                                background=parent.cget("bg") if "bg" in parent.keys() else C_SURFACE,
                                highlightthickness=0)
        self._current_fill = self.fill; self._draw()
        self.canvas.bind("<Enter>", self._on_enter); self.canvas.bind("<Leave>", self._on_leave)
        self.canvas.bind("<ButtonPress-1>", self._on_press); self.canvas.bind("<ButtonRelease-1>", self._on_release)
    def _rounded_rect(self, x, y, w, h, r, fill):
        c = self.canvas
        c.create_rectangle(x+r, y, x+w-r, y+h, outline="", fill=fill)
        c.create_rectangle(x, y+r, x+w, y+h-r, outline="", fill=fill)
        c.create_arc(x, y, x+2*r, y+2*r, start=90, extent=90, outline="", fill=fill)
        c.create_arc(x+w-2*r, y, x+w, y+2*r, start=0, extent=90, outline="", fill=fill)
        c.create_arc(x, y+h-2*r, x+2*r, y+h, start=180, extent=90, outline="", fill=fill)
        c.create_arc(x+w-2*r, y+h-2*r, x+w, y+h, start=270, extent=90, outline="", fill=fill)
    def _draw(self):
        self.canvas.delete("all")
        self._rounded_rect(0, 0, self.w, self.h, self.radius, self._current_fill)
        self.canvas.create_text(self.w//2, self.h//2, text=self.text, fill=self.fg, font=self.font)
    def _on_enter(self,_):
        if not self.disabled: self._current_fill = C_TEAL; self._draw()
    def _on_leave(self,_):
        if not self.disabled: self._current_fill = self.fill; self._draw()
    def _on_press(self,_):
        if not self.disabled: self._current_fill = "#0a6f5c"; self._draw()
    def _on_release(self,_):
        if not self.disabled:
            self._current_fill = C_TEAL; self._draw()
            if callable(self.command): self.command()
    def set_disabled(self, flag: bool):
        self.disabled = flag; self._current_fill = "#bdbdbd" if flag else self.fill; self._draw()
    def pack(self,*a,**k): self.canvas.pack(*a,**k)

def make_card(parent, title: str):
    wrapper = tk.Frame(parent, bg=C_BG)
    head = tk.Frame(wrapper, bg=C_PRIMARY, height=28)
    head.pack(fill=tk.X, side=tk.TOP); head.pack_propagate(False)
    tk.Label(head, text=title, bg=C_PRIMARY, fg="white", font=("Segoe UI Semibold", 11), padx=12)\
        .pack(anchor="w", fill=tk.BOTH)
    body = tk.Frame(wrapper, bg=C_SURFACE, highlightthickness=1, highlightbackground=C_BORDER)
    body.pack(fill=tk.BOTH, expand=True, side=tk.TOP)
    return wrapper, body

class ScrollableFrame(tk.Frame):
    """Container com rolagem vertical (para Passo 1–3)."""
    def __init__(self, parent, bg=C_BG):
        super().__init__(parent, bg=bg)
        self.canvas = tk.Canvas(self, bg=bg, highlightthickness=0)
        self.scrollbar = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.inner = tk.Frame(self.canvas, bg=bg)
        self.inner_id = self.canvas.create_window((0, 0), window=self.inner, anchor="nw")

        self.canvas.configure(yscrollcommand=self.scrollbar.set)
        self.canvas.grid(row=0, column=0, sticky="nsew")
        self.scrollbar.grid(row=0, column=1, sticky="ns")
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)

        self.inner.bind("<Configure>", self._on_inner_configure)
        self.canvas.bind("<Configure>", self._on_canvas_configure)

        # rolagem do mouse
        self.canvas.bind_all("<MouseWheel>", self._on_mousewheel_windows)
        self.canvas.bind_all("<Button-4>", self._on_mousewheel_linux)
        self.canvas.bind_all("<Button-5>", self._on_mousewheel_linux)

    def _on_inner_configure(self, _):
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def _on_canvas_configure(self, event):
        self.canvas.itemconfig(self.inner_id, width=event.width)

    def _on_mousewheel_windows(self, event):
        # apenas se o mouse estiver sobre o canvas:
        if self.winfo_containing(event.x_root, event.y_root) in (self.canvas, self.inner):
            self.canvas.yview_scroll(int(-1*(event.delta/120)), "units")

    def _on_mousewheel_linux(self, event):
        if self.winfo_containing(event.x_root, event.y_root) in (self.canvas, self.inner):
            if event.num == 4:
                self.canvas.yview_scroll(-1, "units")
            elif event.num == 5:
                self.canvas.yview_scroll(1, "units")

class CalibrateWindow(tk.Toplevel):
    def __init__(self, master, doc_path: str, field_title: str = "nome"):
        super().__init__(master)
        self.title(f"Selecionar área de {field_title}")
        self.doc_path = doc_path; self.rect_pdf = None
        self.configure(bg=C_BG); self.geometry("1100x700"); self.minsize(820,560); self.resizable(True, True)

        top = tk.Frame(self, bg=C_BG); top.pack(fill=tk.X, padx=14, pady=(12,6))
        tk.Label(top, text="Zoom:", bg=C_BG, fg=C_WARM).pack(side=tk.LEFT)
        self.zoom_var = tk.DoubleVar(value=1.0); self.zoom = 1.0
        ttk.Scale(top, from_=0.3, to=3.0, orient="horizontal",
                  variable=self.zoom_var, command=lambda *_: self._apply_zoom()).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=8)
        ttk.Button(top, text="–", width=3, command=lambda: self._bump_zoom(-0.1)).pack(side=tk.LEFT, padx=(4,0))
        ttk.Button(top, text="+", width=3, command=lambda: self._bump_zoom(+0.1)).pack(side=tk.LEFT, padx=(4,12))
        ttk.Button(top, text="Largura", command=self._fit_width).pack(side=tk.LEFT, padx=(0,6))
        ttk.Button(top, text="Página", command=self._fit_page).pack(side=tk.LEFT, padx=(0,6))
        ttk.Button(top, text="Reset", command=self._reset_view).pack(side=tk.LEFT)

        tk.Label(self, text="Esquerdo=selecionar • Direito/Meio=arrastar • Roda=zoom",
                 bg=C_BG, fg=C_MUTED).pack(anchor="w", padx=16, pady=(0,4))

        area = tk.Frame(self, bg=C_BG); area.pack(fill=tk.BOTH, expand=True, padx=14, pady=(0,10))
        self.canvas = tk.Canvas(area, background=C_SURFACE, highlightthickness=1, highlightbackground=C_BORDER, cursor="tcross")
        self.hbar = ttk.Scrollbar(area, orient=tk.HORIZONTAL, command=self.canvas.xview)
        self.vbar = ttk.Scrollbar(area, orient=tk.VERTICAL, command=self.canvas.yview)
        self.canvas.configure(xscrollcommand=self.hbar.set, yscrollcommand=self.vbar.set)
        self.canvas.grid(row=0, column=0, sticky="nsew")
        self.vbar.grid(row=0, column=1, sticky="ns")
        self.hbar.grid(row=1, column=0, sticky="ew")
        area.rowconfigure(0, weight=1); area.columnconfigure(0, weight=1)

        try:
            doc = fitz.open(self.doc_path); page = doc[0]
            pix = page.get_pixmap(matrix=fitz.Matrix(1.8,1.8), alpha=False)
            self.img_base = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            self.img_w0, self.img_h0 = self.img_base.width, self.img_base.height
            self.page_w, self.page_h = page.rect.width, page.rect.height
            self.scale_x = self.img_w0 / self.page_w; self.scale_y = self.img_h0 / self.page_h
        finally:
            try: doc.close()
            except Exception: pass

        self.photo = ImageTk.PhotoImage(self.img_base)
        self.img_item = self.canvas.create_image(0, 0, anchor="nw", image=self.photo)
        self.canvas.config(scrollregion=(0,0,self.img_w0,self.img_h0))

        self.sel_rect = None; self.start_x = self.start_y = None
        self.canvas.bind("<ButtonPress-1>", self._on_press)
        self.canvas.bind("<B1-Motion>", self._on_drag)
        self.canvas.bind("<ButtonRelease-1>", self._on_release)
        self.canvas.bind("<ButtonPress-2>", lambda e: self.canvas.scan_mark(e.x, e.y))
        self.canvas.bind("<B2-Motion>",   lambda e: self.canvas.scan_dragto(e.x, e.y, gain=1))
        self.canvas.bind("<ButtonPress-3>", lambda e: self.canvas.scan_mark(e.x, e.y))
        self.canvas.bind("<B3-Motion>",     lambda e: self.canvas.scan_dragto(e.x, e.y, gain=1))
        self.canvas.bind("<MouseWheel>", lambda e: self._bump_zoom(+0.1 if e.delta > 0 else -0.1))

        btns = tk.Frame(self, bg=C_BG); btns.pack(fill=tk.X, padx=14, pady=(6,12))
        ttk.Button(btns, text="Cancelar", command=self.destroy).pack(side=tk.RIGHT)
        ttk.Button(btns, text="Confirmar", command=self._confirm).pack(side=tk.RIGHT, padx=6)

        self.after(120, self._fit_width)

    def _apply_zoom(self):
        z = max(0.3, min(3.0, float(self.zoom_var.get())))
        if abs(z - self.zoom) < 1e-3: return
        self.zoom = z
        w = int(self.img_w0 * self.zoom); h = int(self.img_h0 * self.zoom)
        img = self.img_base.resize((w,h), Image.LANCZOS)
        self.photo = ImageTk.PhotoImage(img)
        self.canvas.itemconfig(self.img_item, image=self.photo)
        self.canvas.config(scrollregion=(0,0,w,h))

    def _bump_zoom(self, dv: float):
        self.zoom_var.set(max(0.3, min(3.0, float(self.zoom_var.get()) + dv))); self._apply_zoom()

    def _fit_width(self):
        self.update_idletasks()
        vis_w = max(200, self.canvas.winfo_width())
        self.zoom_var.set(vis_w / self.img_w0); self._apply_zoom()

    def _fit_page(self):
        self.update_idletasks()
        vis_w = max(200, self.canvas.winfo_width())
        vis_h = max(200, self.canvas.winfo_height())
        z = min(vis_w / self.img_w0, vis_h / self.img_h0)
        self.zoom_var.set(z); self._apply_zoom()

    def _reset_view(self):
        self.zoom_var.set(1.0); self._apply_zoom()
        self.canvas.xview_moveto(0); self.canvas.yview_moveto(0)
        if self.sel_rect: self.canvas.delete(self.sel_rect); self.sel_rect = None

    def _on_press(self, e):
        x = self.canvas.canvasx(e.x); y = self.canvas.canvasy(e.y)
        self.start_x, self.start_y = x, y
        if self.sel_rect: self.canvas.delete(self.sel_rect); self.sel_rect = None
        self.sel_rect = self.canvas.create_rectangle(x, y, x, y, outline=C_TEAL, width=2)

    def _on_drag(self, e):
        if self.start_x is None: return
        x = self.canvas.canvasx(e.x); y = self.canvas.canvasy(e.y)
        self.canvas.coords(self.sel_rect, self.start_x, self.start_y, x, y)

    def _on_release(self, e): pass

    def _confirm(self):
        if not self.sel_rect:
            messagebox.showwarning("Atenção", "Selecione uma área arrastando o mouse."); return
        x0, y0, x1, y1 = self.canvas.coords(self.sel_rect)
        w = self.img_w0 * self.zoom; h = self.img_h0 * self.zoom
        x0 = min(max(0, x0), w); x1 = min(max(0, x1), w)
        y0 = min(max(0, y0), h); y1 = min(max(0, y1), h)
        if abs(x1-x0) < 3 or abs(y1-y0) < 3:
            messagebox.showwarning("Atenção", "Área muito pequena."); return
        ix0, iy0 = x0 / self.zoom, y0 / self.zoom
        ix1, iy1 = x1 / self.zoom, y1 / self.zoom
        pdf_x0, pdf_y0 = ix0 / self.scale_x, iy0 / self.scale_y
        pdf_x1, pdf_y1 = ix1 / self.scale_x, iy1 / self.scale_y
        x0, x1 = sorted([pdf_x0, pdf_x1]); y0, y1 = sorted([pdf_y0, pdf_y1])
        self.rect_pdf = (x0, y0, x1, y1); self.destroy()

# ---------- Dialogo de campos ----------
class TemplateFieldsDialog(tk.Toplevel):
    def __init__(self, master, template_pdf: str):
        super().__init__(master)
        self.title("Campos deste modelo")
        self.configure(bg=C_BG); self.resizable(False, False)
        self.template_pdf = template_pdf
        self.vars: Dict[str, tk.BooleanVar] = {
            FIELD_NAME:  tk.BooleanVar(value=True),
            FIELD_CPF:   tk.BooleanVar(value=False),
            FIELD_DATE:  tk.BooleanVar(value=False),
            FIELD_TURMA: tk.BooleanVar(value=False),
        }
        self.detected: Dict[str, fitz.Rect] = {}
        self.status_labels: Dict[str, tk.Label] = {}

        frm = tk.Frame(self, bg=C_BG, padx=14, pady=12)
        frm.pack(fill=tk.BOTH, expand=True)

        tk.Label(frm, text="Marque o que este PDF usa. Clique em “Detectar automaticamente”\npara localizar as áreas na 1ª página (o app tenta adivinhar).",
                 bg=C_BG, fg=C_WARM, justify="left").grid(row=0, column=0, columnspan=3, sticky="w", pady=(0,10))

        row = 1
        for field in [FIELD_NAME, FIELD_CPF, FIELD_DATE, FIELD_TURMA]:
            cb = ttk.Checkbutton(frm, text=FIELD_TITLES[field], variable=self.vars[field])
            cb.grid(row=row, column=0, sticky="w", pady=4)
            if field == FIELD_NAME: cb.state(["disabled"])
            st = tk.Label(frm, text="—", bg=C_BG, fg=C_MUTED)
            st.grid(row=row, column=1, sticky="w")
            self.status_labels[field] = st
            row += 1

        ttk.Button(frm, text="Detectar automaticamente", command=self.detect_now).grid(row=row, column=0, sticky="w", pady=(8,0))
        ttk.Button(frm, text="OK", command=self.on_ok).grid(row=row, column=2, sticky="e", pady=(8,0))

    def detect_now(self):
        try:
            doc = fitz.open(self.template_pdf)
        except Exception as e:
            messagebox.showerror("PDF", f"Não foi possível abrir o PDF.\n\n{e}"); return
        try:
            self.detected = detect_field_rects_firstpage(doc)
            self.detected_styles = detect_field_styles_firstpage(doc)
            for f, lab in self.status_labels.items():
                if f in self.detected:
                    lab.configure(text="✓ Encontrado", fg="#117a54")
                else:
                    lab.configure(text="× Não encontrado", fg="#a33")
        finally:
            try: doc.close()
            except Exception: pass

    def on_ok(self):
        self.required = [f for f,v in self.vars.items() if v.get()]
        self.destroy()

# ---------- App ----------
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        try:
            import sv_ttk; sv_ttk.set_theme("light")
        except Exception: pass
        self.configure(bg=C_BG)

        # Detecta e ativa "compact mode" para telas menores (≤1366×768)
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        self.compact = (sw <= 1366 or sh <= 768)
        if self.compact:
            try: self.tk.call('tk', 'scaling', 0.9)
            except Exception: pass
            # reduzir dimensões mínimas/globais
            global SIDEBAR_W, PREVIEW_MIN_W, PREVIEW_MIN_H, BAND_HEIGHT, BANNER1_MAX_H
            SIDEBAR_W     = 440
            PREVIEW_MIN_W = 440
            PREVIEW_MIN_H = 260
            BAND_HEIGHT   = 52
            BANNER1_MAX_H = 36

        # queue UI
        self._q: "queue.Queue[Tuple[str, object]]" = queue.Queue()
        self.after(50, self._pump_queue)

        self._brand_after_band = None
        self._brand_after_left = None
        self._banner1_dims = (0, 0)
        self._banner2_dims = (0, 0)

        self.grid_rowconfigure(0, weight=0, minsize=BAND_HEIGHT)
        self.grid_rowconfigure(1, weight=1)
        self.grid_columnconfigure(0, weight=1)

        self.band = tk.Frame(self, bg=C_PRIMARY, height=BAND_HEIGHT)
        self.band.grid(row=0, column=0, sticky="nsew"); self.band.grid_propagate(False)
        self._band_label = tk.Label(self.band, bg=C_PRIMARY)
        self._band_label.pack(side="left", padx=BANNER_SIDE_PAD, pady=8)

        content = tk.Frame(self, bg=C_BG)
        content.grid(row=1, column=0, sticky="nsew")
        content.grid_rowconfigure(0, weight=1)
        content.grid_columnconfigure(0, weight=1)
        content.grid_columnconfigure(1, weight=0, minsize=SIDEBAR_W)

        # ESQUERDA: agora é rolável
        left_scroller = ScrollableFrame(content, bg=C_BG)
        left_scroller.grid(row=0, column=0, sticky="nsew", padx=(20,10), pady=(14,12))
        self.left = left_scroller.inner

        # DIREITA (sidebar)
        self.right = tk.Frame(content, bg=C_BG)
        self.right.grid(row=0, column=1, sticky="nsew", padx=(10,20), pady=(14,12))

        style = ttk.Style()
        style.configure("TLabel", foreground=C_WARM)
        style.configure("TCheckbutton", foreground=C_WARM)
        style.configure("App.Horizontal.TProgressbar", troughcolor=C_SURFACE)

        # Estados
        self._cancel = False
        self.template_pdf = tk.StringVar(value="")
        self.xlsx_path    = tk.StringVar(value="")
        self.font_path    = tk.StringVar(value="")
        self.output_dir   = tk.StringVar(value="")
        self.evento       = tk.StringVar(value="Fisio Summit BR 2025")

        # --- estados por campo ---
        self.adjust_field = tk.StringVar(value=FIELD_NAME)
        self._offset_vars: Dict[str, Dict[str, tk.DoubleVar]] = {
            FIELD_NAME:  {"x": tk.DoubleVar(value=0.0), "y": tk.DoubleVar(value=0.0)},
            FIELD_CPF:   {"x": tk.DoubleVar(value=0.0), "y": tk.DoubleVar(value=0.0)},
            FIELD_DATE:  {"x": tk.DoubleVar(value=0.0), "y": tk.DoubleVar(value=0.0)},
            FIELD_TURMA: {"x": tk.DoubleVar(value=0.0), "y": tk.DoubleVar(value=0.0)},
        }
        self._snap_enabled: Dict[str, tk.BooleanVar] = {f: tk.BooleanVar(value=True) for f in ALL_FIELDS}
        self._snap_tol: Dict[str, tk.DoubleVar] = {f: tk.DoubleVar(value=24.0) for f in ALL_FIELDS}
        self._baseline_tweak: Dict[str, tk.DoubleVar] = {f: tk.DoubleVar(value=0.0) for f in ALL_FIELDS}
        self._font_override: Dict[str, tk.DoubleVar] = {
            FIELD_NAME:  tk.DoubleVar(value=0.0),
            FIELD_CPF:   tk.DoubleVar(value=0.0),
            FIELD_DATE:  tk.DoubleVar(value=0.0),
            FIELD_TURMA: tk.DoubleVar(value=0.0),
        }
        self._use_consistent_size: Dict[str, tk.BooleanVar] = {
            FIELD_NAME: tk.BooleanVar(value=False),
            FIELD_CPF:  tk.BooleanVar(value=False),
            FIELD_DATE: tk.BooleanVar(value=False),
            FIELD_TURMA:tk.BooleanVar(value=False),
        }
        # Negrito por campo (UI)
        self._bold_vars: Dict[str, tk.BooleanVar] = {f: tk.BooleanVar(value=False) for f in ALL_FIELDS}

        self.preview_name = tk.StringVar(value="Seu Nome")
        self.merge_pdf = tk.BooleanVar(value=False)
        self.progress_val = tk.IntVar(value=0)
        self.status_var = tk.StringVar(value="")
        self.default_turma = tk.StringVar(value="")

        self.template_cfg: Optional[TemplateConfig] = None

        self._init_icons()
        self._load_brand_images()

        self._build_left()
        self._build_right()
        self._bind_offset_traces()
        self.refresh_area_status()

        self._pv_after_id = None
        self.band.bind("<Configure>", self._on_band_configure)

        self.bind_all("<Control-o>", lambda e: self.pick_template())
        self.bind_all("<Control-Shift-o>", lambda e: self.pick_xlsx())
        self.bind_all("<Control-f>", lambda e: self.pick_font())
        self.bind_all("<Control-g>", lambda e: self.generate())
        self.bind_all("<F1>", lambda e: self._show_help())
        self.protocol("WM_DELETE_WINDOW", self._on_close)

        try: self.iconbitmap(asset("logo.ico"))
        except Exception: pass

        try: self.state("zoomed")
        except Exception:
            self.geometry(f"{sw}x{sh}+0+0")

        self.after(120, self._apply_brand_images)
        self.after(350, self.refresh_preview)
        self.after(500, self._show_tutorial)

    # ---- queue pump ----
    def _pump_queue(self):
        try:
            while True:
                kind, payload = self._q.get_nowait()
                if kind == "log":
                    self.status_var.set(str(payload).strip()); self.update_idletasks()
                elif kind == "prog":
                    v = int(payload); self.pb["value"] = v; self.progress_val.set(v)
                elif kind == "max":
                    m = int(payload); self.pb.configure(maximum=m); self.pb["value"] = 0; self.progress_val.set(0)
                elif kind == "enable_cancel":
                    self.btn_cancel.set_disabled(not bool(payload))
                elif kind == "enable_generate":
                    self.btn_generate.set_disabled(not bool(payload))
        except queue.Empty:
            pass
        finally:
            self.after(50, self._pump_queue)

    def _init_icons(self): pass

    def _load_brand_images(self):
        self._banner1_img_raw = None; self._banner2_img_raw = None
        try:
            p1 = asset("banner1.png")
            if os.path.isfile(p1): self._banner1_img_raw = Image.open(p1).convert("RGBA")
        except Exception as e:
            self._banner1_img_raw = None; log_warn(f"banner1 load: {e}")
        try:
            p2 = asset("banner2.png")
            if os.path.isfile(p2): self._banner2_img_raw = Image.open(p2).convert("RGBA")
        except Exception as e:
            self._banner2_img_raw = None; log_warn(f"banner2 load: {e}")

    def _on_band_configure(self, _):
        if self._brand_after_band:
            try: self.after_cancel(self._brand_after_band)
            except Exception: pass
        self._brand_after_band = self.after(120, self._apply_brand_images)

    def _apply_brand_images(self):
        if self._banner1_img_raw is not None and hasattr(self, "_band_label"):
            bw = max(0, self.band.winfo_width()); bh = max(0, self.band.winfo_height())
            if bw > 4 and bh > 4:
                iw, ih = self._banner1_img_raw.size
                avail_w = max(1, bw - 2*BANNER_SIDE_PAD)
                avail_h = max(1, min(BANNER1_MAX_H, bh - 16))
                scale = min(avail_w / iw, avail_h / ih, 1.0)
                new_w = max(1, int(iw * scale)); new_h = max(1, int(ih * scale))
                if (new_w, new_h) != self._banner1_dims:
                    new = self._banner1_img_raw.resize((new_w, new_h), Image.LANCZOS)
                    self._banner1_photo = ImageTk.PhotoImage(new)
                    self._band_label.configure(image=self._banner1_photo)
                    self._band_label.image = self._banner1_photo
                    self._banner1_dims = (new_w, new_h)

        if self._banner2_img_raw is not None and hasattr(self, "_banner2_label"):
            # o banner 2 fica na coluna esquerda; o ScrollableFrame ajusta largura
            pass

    # ---------- LEFT ----------
    def _build_left(self):
        # Linha de marca (banner secundário) opcional — removida para economizar altura em telas pequenas
        wrap1, g = make_card(self.left, "Passo 1 — O que você vai usar")
        wrap1.pack(fill=tk.X, padx=0, pady=(0,10))
        for i in range(4): g.columnconfigure(i, weight=1 if i==1 else 0)

        ttk.Label(g, text="Modelo do certificado (PDF):").grid(row=0, column=0, sticky="w", padx=12, pady=(10,4))
        ttk.Entry(g, textvariable=self.template_pdf).grid(row=0, column=1, sticky="ew", padx=6, pady=(10,4))
        ttk.Button(g, text="Escolher...", command=self.pick_template).grid(row=0, column=2, padx=12, pady=(10,4), sticky="e")

        ttk.Label(g, text="Planilha (.XLSX):").grid(row=1, column=0, sticky="w", padx=12, pady=4)
        ttk.Entry(g, textvariable=self.xlsx_path).grid(row=1, column=1, sticky="ew", padx=6, pady=4)
        ttk.Button(g, text="Como preparar XLSX?", command=self._explain_xlsx).grid(row=1, column=2, padx=(6,0), pady=4)
        ttk.Button(g, text="Escolher...", command=self.pick_xlsx).grid(row=1, column=3, padx=12, pady=4, sticky="e")

        ttk.Label(g, text="Fonte (.TTF/.OTF):").grid(row=2, column=0, sticky="w", padx=12, pady=4)
        ttk.Entry(g, textvariable=self.font_path).grid(row=2, column=1, sticky="ew", padx=6, pady=4)
        ttk.Button(g, text="Escolher...", command=self.pick_font).grid(row=2, column=2, padx=12, pady=4, sticky="e")

        ttk.Label(g, text="Nome do evento (aparece no arquivo):").grid(row=3, column=0, sticky="w", padx=12, pady=4)
        self.combo_evento = ttk.Combobox(g, textvariable=self.evento, state="readonly",
                                         values=["Fisio Summit 2025","Liberação Funcional Avançada",
                                                 "Mobilização Neural","RCA360",
                                                 "Congresso RCA 2026"])
        self.combo_evento.grid(row=3, column=1, sticky="ew", padx=6, pady=4)

        ttk.Label(g, text="Pasta onde salvar (opcional):").grid(row=4, column=0, sticky="w", padx=12, pady=(4,12))
        ttk.Entry(g, textvariable=self.output_dir).grid(row=4, column=1, sticky="ew", padx=6, pady=(4,12))
        ttk.Button(g, text="Escolher...", command=self.pick_output_dir).grid(row=4, column=2, padx=12, pady=(4,12), sticky="e")

        # ----- Passo 2 -----
        wrap2, s2 = make_card(self.left, "Passo 2 — Onde os campos entram (1ª página)")
        wrap2.pack(fill=tk.X, padx=0, pady=(0,10))
        for c in range(4): s2.grid_columnconfigure(c, weight=1)

        ttk.Label(s2, text=("Selecione as áreas do Nome/CPF/Data/Turma (ou deixe o automático). "
                            "Os botões ficam ativos apenas para os campos escolhidos no início.")
                 ).grid(row=0, column=0, columnspan=4, sticky="w", padx=12, pady=(10,6))

        self.btn_sel_nome  = ttk.Button(s2, text="Selecionar área do Nome…",  command=lambda: self.calibrate_field(FIELD_NAME))
        self.btn_sel_cpf   = ttk.Button(s2, text="CPF…",   command=lambda: self.calibrate_field(FIELD_CPF))
        self.btn_sel_data  = ttk.Button(s2, text="Data…",  command=lambda: self.calibrate_field(FIELD_DATE))
        self.btn_sel_turma = ttk.Button(s2, text="Turma…", command=lambda: self.calibrate_field(FIELD_TURMA))
        self.btn_sel_nome.grid(row=1, column=0, sticky="ew", padx=12, pady=(0,10))
        self.btn_sel_cpf.grid(row=1, column=1, sticky="ew", padx=6,  pady=(0,10))
        self.btn_sel_data.grid(row=1, column=2, sticky="ew", padx=6,  pady=(0,10))
        self.btn_sel_turma.grid(row=1, column=3, sticky="ew", padx=12, pady=(0,10))

        self.area_status = tk.StringVar(value="Áreas: —")
        ttk.Label(s2, textvariable=self.area_status).grid(row=2, column=0, columnspan=4, sticky="w", padx=12, pady=(0,10))

        # ----- Passo 3 -----
        wrap3, step3 = make_card(self.left, "Passo 3 — Ajustes da pré-visualização")
        wrap3.pack(fill=tk.BOTH, expand=True, padx=0, pady=(0,4))

        step3.columnconfigure(0, weight=1, minsize=420 if self.compact else 460)
        step3.columnconfigure(1, weight=1, minsize=380 if self.compact else 420)

        # Linha topo: campo ativo
        field_row = tk.Frame(step3, bg=C_SURFACE)
        field_row.grid(row=0, column=0, columnspan=2, sticky="we", padx=12, pady=(10,4))
        ttk.Label(field_row, text="Qual campo você quer ajustar agora?").pack(side="left")
        for label,val in [("Nome",FIELD_NAME),("CPF",FIELD_CPF),("Data",FIELD_DATE),("Turma",FIELD_TURMA)]:
            ttk.Radiobutton(field_row, text=label, value=val, variable=self.adjust_field,
                            command=self._on_adjust_field_change).pack(side="left", padx=(8,0))

        # Coluna esquerda
        leftcol = tk.Frame(step3, bg=C_SURFACE)
        leftcol.grid(row=1, column=0, sticky="nsew", padx=(12,6), pady=(6,12))
        leftcol.columnconfigure(0, weight=1)
        leftcol.columnconfigure(1, weight=0)

        # Snap + Negrito (lado a lado)
        self.cb_snap = ttk.Checkbutton(
            leftcol,
            text="Alinhar automaticamente com textos/linhas do PDF (snap)",
            command=self.refresh_preview
        )
        self.cb_snap.grid(row=0, column=0, sticky="w")
        # Negrito para CPF/Turma
        self.chk_bold = ttk.Checkbutton(
            leftcol, text="Negrito", command=self._on_bold_toggle
        )
        self.chk_bold.grid(row=0, column=1, sticky="e", padx=(8,0))

        self.cb_consistent = ttk.Checkbutton(
            leftcol,
            text="Usar o MESMO tamanho de letra para TODOS os nomes (consistência)",
            command=self.refresh_preview
        )
        self.cb_consistent.grid(row=1, column=0, columnspan=2, sticky="w", pady=(6,4))

        # Ajustes finos
        nudges = ttk.LabelFrame(leftcol, text="Ajustes finos deste campo")
        nudges.grid(row=2, column=0, columnspan=2, sticky="nsew", pady=(8,0), ipadx=6, ipady=6, padx=0)
        for i in range(2): nudges.columnconfigure(i, weight=0)

        ttk.Label(nudges, text="Mover para os lados (X, em pt):").grid(row=0, column=0, padx=(8,6), pady=(2,2), sticky="e")
        self.spin_offx = ttk.Spinbox(nudges, from_=-300, to=300, increment=0.5, width=10, command=self._on_nudge_spin)
        self.spin_offx.grid(row=0, column=1, sticky="w", pady=(2,2))

        ttk.Label(nudges, text="Mover para cima/baixo (Y, em pt):").grid(row=1, column=0, padx=(8,6), pady=(2,2), sticky="e")
        self.spin_offy = ttk.Spinbox(nudges, from_=-300, to=300, increment=0.5, width=10, command=self._on_nudge_spin)
        self.spin_offy.grid(row=1, column=1, sticky="w", pady=(2,2))

        # Tamanho de fonte (CPF/Data/Turma) — 0 = automático
        self.font_frame = ttk.LabelFrame(leftcol, text="Tamanho da fonte (CPF / Data / Turma)")
        self.font_frame.grid(row=3, column=0, columnspan=2, sticky="nsew", pady=(8,0), ipadx=6, ipady=6)
        self.font_frame.columnconfigure(0, weight=0)
        self.font_frame.columnconfigure(1, weight=1)
        ttk.Label(self.font_frame, text="Tamanho (pt, 0 = auto):").grid(row=0, column=0, padx=(8,6), sticky="e")
        self.spin_font = ttk.Spinbox(self.font_frame, from_=0.0, to=240.0, increment=0.5, width=10)
        self.spin_font.grid(row=0, column=1, sticky="w")

        # Coluna direita — tolerâncias / linha-base + prévia
        rightcol = tk.Frame(step3, bg=C_SURFACE)
        rightcol.grid(row=1, column=1, sticky="nsew", padx=(6,12), pady=(6,12))
        rightcol.columnconfigure(0, weight=0, minsize=120)
        rightcol.columnconfigure(1, weight=1, minsize=280)

        rowr = 0
        ttk.Label(rightcol, text="Força do “grude” (snap) com o layout (pt):").grid(row=rowr, column=0, sticky="e")
        self.spin_tol = ttk.Spinbox(rightcol, from_=0, to=120, increment=1, width=8, command=self.refresh_preview)
        self.spin_tol.grid(row=rowr, column=1, sticky="w")
        rowr += 1

        ttk.Label(rightcol, text="Subir/Descer as letras (linha-base, em “ems”):").grid(row=rowr, column=0, sticky="e", pady=(8,0))
        self.spin_base = ttk.Spinbox(rightcol, from_=-0.3, to=0.3, increment=0.01, width=8, command=self.refresh_preview)
        self.spin_base.grid(row=rowr, column=1, sticky="w", pady=(8,0))
        rowr += 1

        # Controle de pré-visualização
        pv_ctrl = ttk.LabelFrame(rightcol, text="Pré-visualização (página inteira)")
        pv_ctrl.grid(row=rowr, column=0, columnspan=2, sticky="we", pady=(12,0), ipadx=6, ipady=6)
        pv_ctrl.columnconfigure(0, weight=0, minsize=120)
        pv_ctrl.columnconfigure(1, weight=1, minsize=220 if self.compact else 320)

        ttk.Label(pv_ctrl, text="Nome de teste:").grid(row=0, column=0, sticky="e", padx=(6,4))
        e = ttk.Entry(pv_ctrl, textvariable=self.preview_name, width=28)
        e.grid(row=0, column=1, sticky="we", padx=(0,8))
        e.bind("<KeyRelease>", lambda *_: self.refresh_preview())

        # Botão logo ABAIXO do campo
        ttk.Button(pv_ctrl, text="← Maior nome da planilha", command=self.set_longest_from_xlsx)\
            .grid(row=1, column=1, padx=(0,8), pady=(6,0), sticky="w")

        # wraplength atualizado ao redimensionar
        step3.bind("<Configure>", lambda _e: self._update_wraplengths(leftcol))
        self._sync_controls_for_field()

    def _build_right(self):
        prev_wrap, prev_body = make_card(self.right, "Pré-visualização")
        prev_wrap.pack(fill=tk.BOTH, expand=True, padx=0, pady=(0,12))
        self._pv_wrap = tk.Frame(prev_body, bg=C_SURFACE)
        self._pv_wrap.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)

        self.preview_canvas = tk.Canvas(self._pv_wrap, background=C_SURFACE,
                                        width=PREVIEW_MIN_W, height=PREVIEW_MIN_H,
                                        highlightthickness=1, highlightbackground=C_BORDER)
        self.preview_canvas.pack(fill=tk.BOTH, expand=True)
        self.preview_img_item = None; self._preview_photo = None
        self.preview_canvas.bind("<Map>", lambda e: self.after(80, self.refresh_preview))
        self.preview_canvas.bind("<Configure>", self._on_canvas_configure)

        actions_wrap, actions_body = make_card(self.right, "Ações")
        actions_wrap.pack(fill=tk.X, padx=0, pady=(0,0))
        pad = 12
        self.btn_generate = RoundedButton(actions_body, "GERAR CERTIFICADOS", command=self.generate)
        self.btn_generate.pack(padx=pad, pady=(pad,8), anchor="n")
        self.btn_cancel = RoundedButton(actions_body, "Cancelar", command=self.cancel_generation, fill="#e0e0e0", fg="#333333")
        self.btn_cancel.set_disabled(True)
        self.btn_cancel.pack(padx=pad, pady=(0,10), anchor="n")
        ttk.Checkbutton(actions_body, text="Juntar tudo em 1 PDF (opcional)", variable=self.merge_pdf)\
            .pack(anchor="w", padx=pad, pady=(0,10))
        ttk.Label(actions_body, text="Progresso:").pack(anchor="w", padx=pad, pady=(0,4))
        self.pb = ttk.Progressbar(actions_body, mode="determinate", maximum=100, variable=self.progress_val, style="App.Horizontal.TProgressbar")
        self.pb.pack(fill=tk.X, padx=pad, pady=(0,6))
        ttk.Label(actions_body, textvariable=self.status_var, foreground=C_MUTED, wraplength=SIDEBAR_W-40, justify="left")\
            .pack(anchor="w", padx=pad, pady=(0,10))
        ttk.Button(actions_body, text="Abrir pasta de saída", command=lambda: open_folder(self.output_dir.get().strip() or os.getcwd()))\
            .pack(fill=tk.X, padx=pad, pady=(0,pad))

    # ---------- Helpers ----------
    def _update_wraplengths(self, leftcol: tk.Frame):
        try:
            w = max(240, leftcol.winfo_width() - 24)
            for cb in (self.cb_snap, self.cb_consistent):
                cb.configure(wraplength=w)
        except Exception as e:
            log_warn(f"_update_wraplengths: {e}")

    def _bind_offset_traces(self):
        def bind_spinbox(spin: ttk.Spinbox):
            try:
                spin.bind("<KeyRelease>", lambda _e: self._on_nudge_spin())
                spin.bind("<FocusOut>", lambda _e: self._on_nudge_spin())
            except Exception: pass
        if hasattr(self, "spin_offx"): bind_spinbox(self.spin_offx)
        if hasattr(self, "spin_offy"): bind_spinbox(self.spin_offy)

        for f in ALL_FIELDS:
            self._offset_vars[f]["x"].trace_add("write", lambda *_: self.refresh_preview())
            self._offset_vars[f]["y"].trace_add("write", lambda *_: self.refresh_preview())
            self._snap_enabled[f].trace_add("write", lambda *_: self.refresh_preview())
            self._snap_tol[f].trace_add("write", lambda *_: self.refresh_preview())
            self._baseline_tweak[f].trace_add("write", lambda *_: self.refresh_preview())
        self._use_consistent_size[FIELD_NAME].trace_add("write", lambda *_: self.refresh_preview())
        for f in ALL_FIELDS:
            self._font_override[f].trace_add("write", lambda *_: self.refresh_preview())
            self._bold_vars[f].trace_add("write", lambda *_: self.refresh_preview())

    def _on_canvas_configure(self, _):
        if getattr(self, "_pv_after_id", None):
            try: self.after_cancel(self._pv_after_id)
            except Exception: pass
        self._pv_after_id = self.after(120, self.refresh_preview)

    def log_insert(self, text: str):
        self.status_var.set(text.strip()); self.update_idletasks()

    def refresh_area_status(self):
        parts = []
        if self.template_cfg:
            for f in ALL_FIELDS:
                mark = "✓" if (self.template_cfg and self.template_cfg.fields.get(f)) else "—"
                req  = " (usa)" if (self.template_cfg and f in self.template_cfg.required) else ""
                parts.append(f"{FIELD_TITLES[f]}: {mark}{req}")
        else:
            parts.append("nenhum modelo carregado")
        self.area_status.set(" • ".join(parts))
        self._update_field_buttons_state()

    def _update_field_buttons_state(self):
        active = set(self.template_cfg.required) if self.template_cfg else {FIELD_NAME}
        for f, btn in [(FIELD_NAME, self.btn_sel_nome),
                       (FIELD_CPF, self.btn_sel_cpf),
                       (FIELD_DATE, self.btn_sel_data),
                       (FIELD_TURMA, self.btn_sel_turma)]:
            state = "normal" if f in active else "disabled"
            try: btn.state([state])
            except Exception: btn.configure(state=state)

    def _apply_styles_from_cfg(self):
        for f in ALL_FIELDS:
            val = False
            if self.template_cfg and f in self.template_cfg.fields:
                st = self.template_cfg.fields[f].style or {}
                val = bool(st.get("bold", False))
            self._bold_vars[f].set(val)

    # ---------- Template & fields ----------
    def pick_template(self):
        p = filedialog.askopenfilename(title="Escolha o PDF do modelo", filetypes=[("PDF","*.pdf")])
        if not p: return
        self.template_pdf.set(p)

        cfg = load_template_config(p) or TemplateConfig(required=[FIELD_NAME], fields={}, version=3)
        dlg = TemplateFieldsDialog(self, p)
        self.wait_window(dlg)
        required = getattr(dlg, "required", [FIELD_NAME])

        try:
            doc = fitz.open(p)
            auto_rects = detect_field_rects_firstpage(doc)
            auto_styles = detect_field_styles_firstpage(doc)
        except Exception as e:
            messagebox.showerror("PDF", f"Não foi possível abrir o PDF.\n\n{e}"); return
        finally:
            try: doc.close()
            except Exception: pass

        fields: Dict[str, FieldCalibration] = dict(cfg.fields)
        for f in required:
            if f in auto_rects:
                r = auto_rects[f]
                style = auto_styles.get(f, {}) if isinstance(auto_styles, dict) else {}
                fields[f] = FieldCalibration(page_index=0, rect=(r.x0,r.y0,r.x1,r.y1), style=style)

        missing = [f for f in required if f not in fields]
        for f in missing:
            if messagebox.askyesno("Calibrar", f"Não encontrei área para {FIELD_TITLES[f]}. Selecionar manualmente agora?"):
                win = CalibrateWindow(self, p, field_title=FIELD_TITLES[f])
                self.wait_window(win)
                if win.rect_pdf:
                    fields[f] = FieldCalibration(page_index=0, rect=win.rect_pdf, style={})

        self.template_cfg = TemplateConfig(required=required, fields=fields, version=3)
        save_template_config(p, self.template_cfg)
        self._apply_styles_from_cfg()

        self.refresh_area_status()
        self.refresh_preview()

    def calibrate_field(self, field: str):
        p = self.template_pdf.get().strip()
        if not p:
            messagebox.showwarning("Atenção","Escolha primeiro o PDF do modelo."); return
        win = CalibrateWindow(self, p, field_title=FIELD_TITLES.get(field, field))
        self.wait_window(win)
        if not win.rect_pdf:
            self.log_insert("Seleção cancelada."); return
        if not self.template_cfg:
            self.template_cfg = TemplateConfig(required=[FIELD_NAME], fields={}, version=3)
        prev_style = (self.template_cfg.fields.get(field).style if self.template_cfg.fields.get(field) else {})
        self.template_cfg.fields[field] = FieldCalibration(page_index=0, rect=win.rect_pdf, style=prev_style or {})
        if field not in self.template_cfg.required:
            self.template_cfg.required.append(field)
        save_template_config(p, self.template_cfg)
        self._apply_styles_from_cfg()
        self.refresh_area_status()
        self.refresh_preview()

    # ---------- XLSX / Fonte / Pasta ----------
    def pick_xlsx(self):
        p = filedialog.askopenfilename(title="Escolha a planilha (.XLSX)", filetypes=[("Excel","*.xlsx")])
        if p:
            self.xlsx_path.set(p)
            try:
                recs = read_records_from_xlsx(p)
                if recs and (not self.preview_name.get() or self.preview_name.get()=="Seu Nome"):
                    self.preview_name.set(recs[0].get(FIELD_NAME,""))
            except Exception as e:
                log_warn(f"pick_xlsx: {e}")
            self.refresh_preview()

    def pick_font(self):
        p = filedialog.askopenfilename(title="Escolha a fonte (.TTF/.OTF)",
                                       filetypes=[("Fontes TrueType/OpenType","*.ttf *.otf"),("Todos","*.*")])
        if p:
            self.font_path.set(p); self.refresh_preview()

    def pick_output_dir(self):
        d = filedialog.askdirectory(title="Escolha a pasta onde salvar")
        if d: self.output_dir.set(d)

    def _explain_xlsx(self):
        messagebox.showinfo(
            "Como preparar o .XLSX",
            "Deixe a planilha ativa assim:\n\n"
            "A: Nome (obrigatório)\n"
            "B: CPF (opcional)\n"
            "C: Turma (opcional)\n"
            "D: Data (opcional; se vazio, usamos a data de hoje)\n\n"
            "Sem cabeçalho — comece na linha 1."
        )

    def set_longest_from_xlsx(self):
        p = self.xlsx_path.get().strip()
        if not (p and os.path.isfile(p)):
            messagebox.showinfo("XLSX", "Escolha primeiro a planilha (.XLSX)."); return
        try:
            recs = read_records_from_xlsx(p)
            names = [r.get(FIELD_NAME,"") for r in recs if r.get(FIELD_NAME)]
            if not names: raise ValueError("Planilha vazia.")
            longest = max(names, key=lambda s: len(s or ""))
            self.preview_name.set(longest)
            self.refresh_preview()
            self.log_insert(f"Prévia: maior nome da planilha → “{longest}”.")
        except Exception as e:
            messagebox.showerror("XLSX", f"Falha ao ler XLSX: {e}")

    # ---------- Sincroniza UI de Passo 3 ----------
    def _sync_controls_for_field(self):
        f = self.adjust_field.get()
        self.cb_snap.configure(variable=self._snap_enabled[f])
        self.spin_tol.configure(textvariable=self._snap_tol[f])
        self.spin_base.configure(textvariable=self._baseline_tweak[f])
        self.spin_offx.configure(textvariable=self._offset_vars[f]["x"])
        self.spin_offy.configure(textvariable=self._offset_vars[f]["y"])

        # Negrito: habilita para CPF/TURMA; desabilita para Nome/Data
        self.chk_bold.configure(variable=self._bold_vars[f])
        if f in (FIELD_CPF, FIELD_TURMA):
            try: self.chk_bold.state(["!disabled"])
            except Exception: self.chk_bold.configure(state="normal")
        else:
            try: self.chk_bold.state(["disabled"])
            except Exception: self.chk_bold.configure(state="disabled")

        if f == FIELD_NAME:
            self.cb_consistent.grid()
            self.cb_consistent.configure(variable=self._use_consistent_size[FIELD_NAME])
        else:
            self.cb_consistent.grid_remove()

        if f in (FIELD_CPF, FIELD_DATE, FIELD_TURMA):
            self.font_frame.grid()
            self.spin_font.configure(textvariable=self._font_override[f])
        else:
            self.font_frame.grid_remove()

        self.refresh_preview()

    def _on_bold_toggle(self):
        f = self.adjust_field.get()
        if f not in (FIELD_CPF, FIELD_TURMA): return
        if not self.template_cfg: return
        if f not in self.template_cfg.fields: return
        st = dict(self.template_cfg.fields[f].style or {})
        st["bold"] = bool(self._bold_vars[f].get())
        self.template_cfg.fields[f].style = st
        tpl = self.template_pdf.get().strip()
        if tpl: save_template_config(tpl, self.template_cfg)
        self.refresh_preview()

    def _on_adjust_field_change(self):
        self._sync_controls_for_field()

    # ---------- Prévia ----------
    def _get_fields_rects(self) -> Dict[str, fitz.Rect]:
        p = self.template_pdf.get().strip()
        rects: Dict[str, fitz.Rect] = {}
        if not (p and os.path.isfile(p)): return rects
        cfg = self.template_cfg or load_template_config(p)
        if cfg:
            for f, fc in cfg.fields.items():
                if fc.page_index == 0: rects[f] = fitz.Rect(*fc.rect)
        if FIELD_NAME not in rects:
            try:
                doc = fitz.open(p); rr = compute_auto_area_firstpage(doc); doc.close()
                rects[FIELD_NAME] = rr
            except Exception as e:
                log_warn(f"_get_fields_rects: {e}")
        return rects

    def _effective_rect(self, page: fitz.Page, field: str, rect_raw: fitz.Rect) -> fitz.Rect:
        offx = float(self._offset_vars[field]["x"].get())
        offy = float(self._offset_vars[field]["y"].get())
        return compute_snapped_rect(
            page, rect_raw,
            bool(self._snap_enabled[field].get()),
            float(self._snap_tol[field].get()),
            offx, offy
        )

    def _on_nudge_spin(self): self.refresh_preview()

    def _place_preview_image(self, photo: ImageTk.PhotoImage):
        cw = max(1, self.preview_canvas.winfo_width()); ch = max(1, self.preview_canvas.winfo_height())
        iw, ih = photo.width(), photo.height()
        cx, cy = cw // 2, ch // 2
        x0, y0 = cx - iw // 2, cy - ih // 2
        x1, y1 = x0 + iw, y0 + ih
        self.preview_canvas.delete("all")
        self.preview_canvas.create_rectangle(x0-6, y0-6, x1+6, y1+6, fill="#e9f2ef", outline="")
        self.preview_canvas.create_rectangle(x0-1, y0-1, x1+1, y1+1, outline=C_BORDER, width=1)
        self.preview_img_item = self.preview_canvas.create_image(cx, cy, image=photo, anchor="center")
        self._preview_photo = photo
        self.preview_canvas.image = photo

    def _draw_placeholder(self, msg="Sem prévia"):
        cw = max(1, self.preview_canvas.winfo_width()); ch = max(1, self.preview_canvas.winfo_height())
        self.preview_canvas.delete("all"); pad = 24
        self.preview_canvas.create_rectangle(0, 0, cw, ch, fill=C_SURFACE, outline=C_BORDER)
        self.preview_canvas.create_rectangle(pad, pad, cw-pad, ch-pad, outline=C_BORDER, dash=(4,3))
        self.preview_canvas.create_text(cw // 2, ch // 2, text=msg, fill=C_MUTED, font=("Segoe UI", 11))
        self.preview_img_item = None; self.preview_canvas.image = None

    def _render_fullpage_image(self, page: fitz.Page, box_w: int, box_h: int) -> ImageTk.PhotoImage:
        src_w, src_h = page.rect.width, page.rect.height
        box_w = max(PREVIEW_MIN_W, int(box_w)); box_h = max(PREVIEW_MIN_H, int(box_h))
        s = min(box_w / src_w, box_h / src_h)
        max_pix = 2400
        if src_w * s > max_pix or src_h * s > max_pix:
            s = min(max_pix / src_w, max_pix / src_h)
        pix = page.get_pixmap(matrix=fitz.Matrix(s, s), alpha=False)
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        return ImageTk.PhotoImage(img)

    def refresh_preview(self, *_):
        if not hasattr(self, "preview_canvas"): return
        if getattr(self, "_in_refresh", False): return
        self._in_refresh = True
        try:
            self.update_idletasks()
            cw = self.preview_canvas.winfo_width(); ch = self.preview_canvas.winfo_height()
            if cw <= 10 or ch <= 10:
                self.after(120, self.refresh_preview); return

            tpl = self.template_pdf.get().strip()
            if not (tpl and os.path.isfile(tpl)):
                self._draw_placeholder("Carregue um PDF do modelo para ver a prévia"); return

            fields_rects = self._get_fields_rects()
            if FIELD_NAME not in fields_rects:
                self._draw_placeholder("Defina a área do Nome (automática/seleção manual)"); return

            box_w = max(PREVIEW_MIN_W, cw); box_h = max(PREVIEW_MIN_H, ch)
            font_path = self.font_path.get().strip()

            try:
                base_doc = fitz.open(tpl); page0 = base_doc[0]
            except Exception as e:
                self._draw_placeholder(f"Erro lendo PDF: {e}"); return

            preview_texts = {
                FIELD_NAME: (self.preview_name.get().strip() or "Seu Nome"),
                FIELD_CPF:  "111.222.333-44",
                FIELD_DATE: datetime.date.today().strftime("%d/%m/%Y"),
                FIELD_TURMA:"3",
            }

            fontsize_common = None
            if font_path and os.path.isfile(font_path) and self._use_consistent_size[FIELD_NAME].get():
                try:
                    recs = read_records_from_xlsx(self.xlsx_path.get()) if (self.xlsx_path.get() and os.path.isfile(self.xlsx_path.get())) else []
                    names = [r.get(FIELD_NAME,"") for r in recs if r.get(FIELD_NAME)]
                    if len(names) > 400: names = names[:400]
                    rect = self._effective_rect(page0, FIELD_NAME, fields_rects[FIELD_NAME])
                    fontsize_common = common_font_size_for_all(names or [preview_texts[FIELD_NAME]], rect, font_path)
                except Exception as e:
                    log_warn(f"fontsize_common: {e}")

            colors: Dict[str, Tuple[float,float,float]] = {}
            for f, rr in fields_rects.items():
                eff = self._effective_rect(page0, f, rr)
                colors[f] = pick_text_color(page0, eff)

            cfg_loaded = (self.template_cfg or load_template_config(tpl) or TemplateConfig(required=[FIELD_NAME], fields={}, version=3))
            field_styles: Dict[str, Dict[str,bool]] = {}
            for f in ALL_FIELDS:
                st = {}
                if f in getattr(cfg_loaded, "fields", {}):
                    try: st.update(cfg_loaded.fields[f].style or {})
                    except Exception: pass
                st["bold"] = bool(self._bold_vars.get(f, tk.BooleanVar(value=False)).get())
                field_styles[f] = st

            try:
                tmp = fitz.open(tpl); p = tmp[0]

                if font_path and os.path.isfile(font_path):
                    # Nome
                    rname = self._effective_rect(p, FIELD_NAME, fields_rects[FIELD_NAME])
                    text = preview_texts[FIELD_NAME]
                    fs = fontsize_common if fontsize_common is not None else autosize_font_to_rect(text, rname, font_path)
                    draw_text_centered(p, rname, text, font_path, fs,
                                       baseline_tweak_ems=float(self._baseline_tweak[FIELD_NAME].get()),
                                       color=colors.get(FIELD_NAME,(0,0,0)),
                                       bold=False)

                    # Outros campos (se requeridos)
                    for f in [FIELD_CPF, FIELD_DATE, FIELD_TURMA]:
                        if f in cfg_loaded.required and f in fields_rects:
                            rr = fields_rects[f]
                            r_eff = self._effective_rect(p, f, rr)
                            try: override = float(self._font_override[f].get())
                            except Exception: override = 0.0
                            fs2 = override if (override and override > 0.0) else autosize_font_to_rect(preview_texts[f], r_eff, font_path)
                            draw_text_left(p, r_eff, preview_texts[f], font_path, fs2,
                                           baseline_tweak_ems=float(self._baseline_tweak[f].get()),
                                           color=colors.get(f,(0,0,0)),
                                           bold=bool(field_styles.get(f,{}).get("bold", False)))

                photo = self._render_fullpage_image(p, box_w, box_h)
                self._place_preview_image(photo)
            except Exception as e:
                self._draw_placeholder(f"Não foi possível gerar a prévia: {e}")
            finally:
                try: tmp.close()
                except Exception: pass
                try: base_doc.close()
                except Exception: pass
        finally:
            self._in_refresh = False

    # ---------- Geração ----------
    def cancel_generation(self):
        self._cancel = True; self._q.put(("log", "Cancelando… aguarde a etapa atual."))

    def _precheck_required_vs_xlsx(self, required: List[str]) -> bool:
        p = self.xlsx_path.get().strip()
        if not (p and os.path.isfile(p)):
            messagebox.showwarning("Atenção","Escolha a planilha (.XLSX)."); return False
        try:
            recs = read_records_from_xlsx(p)
        except Exception as e:
            messagebox.showerror("XLSX", f"Falha ao ler XLSX: {e}"); return False
        if not recs:
            messagebox.showwarning("Atenção","Nenhum nome encontrado na planilha."); return False

        if FIELD_CPF in required and not any((r.get(FIELD_CPF,"").strip() for r in recs)):
            messagebox.showerror("XLSX","O modelo exige CPF, mas a coluna B está vazia."); return False

        if FIELD_TURMA in required and not any((r.get(FIELD_TURMA,"").strip() for r in recs)):
            val = simpledialog.askstring("Turma", "A planilha não tem TURMA. Informe um valor padrão para todos:", initialvalue="")
            if val is None: return False
            self.default_turma.set(val.strip())
        return True

    def generate(self):
        tpl = self.template_pdf.get().strip()
        font_path = self.font_path.get().strip()
        evento = self.evento.get().strip()

        if not tpl or not os.path.isfile(tpl):
            messagebox.showwarning("Atenção","Escolha o PDF do modelo."); return
        try:
            ensure_font_file(font_path)
        except Exception as e:
            messagebox.showwarning("Fonte", str(e)); return
        if not evento:
            messagebox.showwarning("Atenção","Informe o nome do evento."); return

        self.template_cfg = self.template_cfg or load_template_config(tpl) or TemplateConfig(required=[FIELD_NAME], fields={}, version=3)
        required = self.template_cfg.required or [FIELD_NAME]
        if not self._precheck_required_vs_xlsx(required): return

        self._cancel = False
        self.btn_generate.set_disabled(True)
        self.btn_cancel.set_disabled(False)
        self.pb.configure(maximum=100); self.pb["value"] = 0; self.progress_val.set(0)
        self._q.put(("log","Preparando…"))

        threading.Thread(target=self._generate_worker, daemon=True).start()

    def _generate_worker(self):
        try:
            tpl = self.template_pdf.get().strip()
            xlsxp = self.xlsx_path.get().strip()
            font_path = self.font_path.get().strip()
            evento = self.evento.get().strip()
            outdir = self.output_dir.get().strip()
            cfg = self.template_cfg or load_template_config(tpl) or TemplateConfig(required=[FIELD_NAME], fields={}, version=3)

            try:
                recs = read_records_from_xlsx(xlsxp)
            except Exception as e:
                self.after(0, lambda: messagebox.showerror("Erro", f"Falha ao ler XLSX: {e}"))
                self._q.put(("enable_generate", True)); self._q.put(("enable_cancel", False)); return
            if not recs:
                self.after(0, lambda: messagebox.showwarning("Atenção","Nenhum nome encontrado na planilha."))
                self._q.put(("enable_generate", True)); self._q.put(("enable_cancel", False)); return

            fields_rects_raw: Dict[str, fitz.Rect] = {}
            for f, fc in cfg.fields.items():
                if fc.page_index == 0: fields_rects_raw[f] = fitz.Rect(*fc.rect)
            if FIELD_NAME not in fields_rects_raw:
                try:
                    doc = fitz.open(tpl); rr = compute_auto_area_firstpage(doc); doc.close()
                    fields_rects_raw[FIELD_NAME] = rr
                except Exception:
                    self._q.put(("log","Não foi possível definir a área do Nome."))
                    self._q.put(("enable_generate", True)); self._q.put(("enable_cancel", False)); return

            if not outdir:
                base = os.path.dirname(tpl) or os.getcwd()
                outdir = os.path.join(base, f"Certificados - {sanitize_filename(evento)} - {timestamp()}")
            os.makedirs(outdir, exist_ok=True)

            bold_snapshot = {f: bool(self._bold_vars[f].get()) for f in ALL_FIELDS}

            try:
                doc_tmp = fitz.open(tpl); page_tmp = doc_tmp[0]
                rect_name = self._effective_rect(page_tmp, FIELD_NAME, fields_rects_raw[FIELD_NAME])
                names = [r.get(FIELD_NAME,"") for r in recs if r.get(FIELD_NAME)]
                if self._use_consistent_size[FIELD_NAME].get() and names:
                    names_for_size = names
                    if len(names_for_size) > 800:
                        step = max(1, len(names_for_size)//800)
                        names_for_size = names_for_size[::step]
                    fontsize_common = common_font_size_for_all(names_for_size, rect_name, font_path)
                else:
                    fontsize_common = None
                colors: Dict[str, Tuple[float,float,float]] = {}
                for f, rr in fields_rects_raw.items():
                    eff = self._effective_rect(page_tmp, f, rr)
                    colors[f] = pick_text_color(page_tmp, eff)
                field_styles: Dict[str, Dict[str,bool]] = {}
                for f in ALL_FIELDS:
                    st = {}
                    if f in cfg.fields:
                        try: st.update(cfg.fields[f].style or {})
                        except Exception: pass
                    st["bold"] = bool(bold_snapshot.get(f, st.get("bold", False)))
                    field_styles[f] = st
            finally:
                try: doc_tmp.close()
                except Exception: pass

            total = len(recs)
            self._q.put(("max", total))
            self._q.put(("log", f"Gerando {total} certificados…"))

            ok = fail = 0
            merged_doc = fitz.Document() if self.merge_pdf.get() else None
            default_turma = self.default_turma.get().strip()

            for idx, rec in enumerate(recs, start=1):
                if self._cancel:
                    self._q.put(("log", "⚠️ Processo cancelado.")); break
                try:
                    doc = fitz.open(tpl)

                    name = rec.get(FIELD_NAME,"").strip()
                    if not name: raise ValueError("Registro sem nome.")

                    date_str = (rec.get(FIELD_DATE,"").strip() or datetime.date.today().strftime("%d/%m/%Y"))
                    turma_str = rec.get(FIELD_TURMA,"").strip() or default_turma
                    cpf_str = rec.get(FIELD_CPF,"").strip()
                    if FIELD_CPF in cfg.required:
                        if not cpf_str: raise ValueError("CPF requerido, mas ausente na planilha.")
                        cpf_str = format_cpf(cpf_str)

                    rect_name = self._effective_rect(doc[0], FIELD_NAME, fields_rects_raw[FIELD_NAME])
                    fs = fontsize_common if ('fontsize_common' in locals() and fontsize_common is not None) else autosize_font_to_rect(name, rect_name, font_path)
                    draw_text_centered(doc[0], rect_name, name, font_path, fs,
                                       baseline_tweak_ems=float(self._baseline_tweak[FIELD_NAME].get()),
                                       color=colors.get(FIELD_NAME,(0,0,0)),
                                       bold=False)

                    for f, text in [(FIELD_CPF, cpf_str), (FIELD_DATE, date_str), (FIELD_TURMA, turma_str)]:
                        if f in cfg.required and f in fields_rects_raw and text:
                            rect_eff = self._effective_rect(doc[0], f, fields_rects_raw[f])
                            try: override = float(self._font_override[f].get())
                            except Exception: override = 0.0
                            fs2 = override if (override and override > 0.0) else autosize_font_to_rect(text, rect_eff, font_path)
                            draw_text_left(doc[0], rect_eff, text, font_path, fs2,
                                           baseline_tweak_ems=float(self._baseline_tweak[f].get()),
                                           color=colors.get(f,(0,0,0)),
                                           bold=bool(field_styles.get(f,{}).get("bold", False)))

                    outname = f"Certificado - {sanitize_filename(evento)} - {sanitize_filename(name)}.pdf"
                    outpath = unique_path(os.path.join(outdir, outname))
                    doc.save(outpath)

                    if merged_doc is not None:
                        try: merged_doc.insert_pdf(doc)
                        except Exception as ie:
                            self._q.put(("log", f"(!) Falha ao montar PDF único: {ie}"))

                    doc.close()
                    ok += 1
                    self._q.put(("log", f"[{idx}/{total}] OK: {name}"))
                except Exception as e:
                    fail += 1
                    self._q.put(("log", f"[{idx}/{total}] ERRO: {rec.get(FIELD_NAME,'(sem nome)')} → {e}"))
                self._q.put(("prog", idx))

            if (merged_doc is not None) and ok > 0:
                try:
                    merged_path = unique_path(os.path.join(outdir, f"Lote — {sanitize_filename(evento)}.pdf"))
                    merged_doc.save(merged_path)
                    self._q.put(("log", f"PDF único criado: {merged_path}"))
                except Exception as e:
                    self._q.put(("log", f"Falha ao criar PDF único: {e}"))
                finally:
                    try: merged_doc.close()
                    except Exception: pass

            self._q.put(("enable_generate", True)); self._q.put(("enable_cancel", False))
            if not self._cancel:
                self._q.put(("log", f"Concluído: {ok} ok, {fail} falhas."))
                self.after(0, lambda: messagebox.showinfo("Pronto", f"Concluído: {ok} ok, {fail} falhas.\nPasta: {outdir}"))
                self.after(0, lambda: open_folder(outdir))
            else:
                self._q.put(("log", f"Interrompido: {ok} ok, {fail} falhas. Parcial salvo em: {outdir}"))
            self.after(0, self.refresh_preview)
        except Exception as e:
            log_err("generate_worker crashed", e)
            self._q.put(("enable_generate", True)); self._q.put(("enable_cancel", False))
            self._q.put(("log", f"Erro inesperado: {e}"))

    def _show_help(self):
        messagebox.showinfo("Atalhos",
            "• Ctrl+O: Abrir PDF do modelo\n"
            "• Ctrl+Shift+O: Abrir planilha (.XLSX)\n"
            "• Ctrl+F: Escolher Fonte (.TTF/.OTF)\n"
            "• Ctrl+G: Gerar Certificados\n"
            "• Na seleção de área: zoom com slider/+/–/roda • arrastar com botão direito/meio.")

    def _show_tutorial(self):
        messagebox.showinfo(
            "Guia rápido",
            "1) Modelo (PDF): Clique em “Escolher...” e selecione o certificado.\n"
            "2) Planilha (.XLSX): A=Nome, B=CPF (opcional), C=Turma (opcional), D=Data (opcional).\n"
            "3) Fonte (.TTF/.OTF): A fonte usada para escrever os campos.\n"
            "4) Campos: Ao abrir o PDF, marque quais campos este modelo usa (o app tenta localizar na 1ª página).\n"
            "   • Se algo não bater, use o Passo 2 para selecionar manualmente.\n"
            "   • Para TURMA/CPF, dá para ligar “Negrito”.\n"
            "5) Pré-visualização (Passo 3): Escolha o CAMPO ATIVO. Ajustes valem para esse campo.\n"
            "   • Snap (alinhar) “gruda” no layout — ajuste a força se precisar.\n"
            "   • Linha-base faz micro-ajuste vertical.\n"
            "6) Gerar: Clique em “GERAR CERTIFICADOS”. “Juntar tudo em 1 PDF” é opcional."
        )

    def _on_close(self): self.destroy()

if __name__ == "__main__":
    App().mainloop()
